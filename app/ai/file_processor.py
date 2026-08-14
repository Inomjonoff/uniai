"""
Document and file processing module.
Extracts text from PDF, DOCX, TXT, CSV, XLSX and chunks it for RAG ingestion.
"""
import io
import csv
from typing import List, Dict, Any, Optional
from app.ai.gemini_client import gemini_client
from app.utils.logger import logger

# Optional document parsing libraries
try:
    import pypdf
    PYPDF_AVAILABLE = True
except ImportError:
    PYPDF_AVAILABLE = False

try:
    import docx
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class FileProcessor:
    def __init__(self, client=gemini_client):
        self.client = client

    def extract_text(self, file_bytes: bytes, file_name: str) -> str:
        """Extracts text based on file extension."""
        ext = file_name.lower().split(".")[-1] if "." in file_name else ""

        try:
            if ext == "pdf":
                return self._extract_pdf(file_bytes)
            elif ext in ("docx", "doc"):
                return self._extract_docx(file_bytes)
            elif ext in ("txt", "md", "log", "json", "yaml", "yml", "sql", "py", "sh"):
                return file_bytes.decode("utf-8", errors="ignore")
            elif ext == "csv":
                return self._extract_csv(file_bytes)
            elif ext in ("xlsx", "xls"):
                return self._extract_excel(file_bytes)
            else:
                return file_bytes.decode("utf-8", errors="ignore")
        except Exception as e:
            logger.error(f"Error extracting text from {file_name}: {e}", exc_info=True)
            return ""

    def _extract_pdf(self, file_bytes: bytes) -> str:
        if not PYPDF_AVAILABLE:
            logger.warning("pypdf is not installed.")
            return ""
        reader = pypdf.PdfReader(io.BytesIO(file_bytes))
        pages_text = []
        for i, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            if text.strip():
                pages_text.append(f"--- Page {i+1} ---\n{text}")
        return "\n\n".join(pages_text)

    def _extract_docx(self, file_bytes: bytes) -> str:
        if not DOCX_AVAILABLE:
            logger.warning("python-docx is not installed.")
            return ""
        doc = docx.Document(io.BytesIO(file_bytes))
        return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])

    def _extract_csv(self, file_bytes: bytes) -> str:
        text_stream = io.StringIO(file_bytes.decode("utf-8", errors="ignore"))
        reader = csv.reader(text_stream)
        rows = [" | ".join(row) for row in reader if any(cell.strip() for cell in row)]
        return "\n".join(rows[:500])  # limit to top 500 rows

    def _extract_excel(self, file_bytes: bytes) -> str:
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl is not installed.")
            return ""
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheets_text = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = []
            for row in sheet.iter_rows(values_only=True):
                non_empty = [str(cell) for cell in row if cell is not None and str(cell).strip()]
                if non_empty:
                    rows.append(" | ".join(non_empty))
            if rows:
                sheets_text.append(f"Sheet: {sheet_name}\n" + "\n".join(rows[:200]))
        return "\n\n".join(sheets_text)

    def chunk_text(self, text: str, chunk_size: int = 1200, overlap: int = 150) -> List[str]:
        """Splits long document text into overlapping chunks."""
        if not text or len(text) <= chunk_size:
            return [text] if text.strip() else []

        chunks = []
        start = 0
        while start < len(text):
            end = start + chunk_size
            chunk = text[start:end]
            # Try to break on a newline or paragraph boundary if possible
            if end < len(text):
                last_newline = chunk.rfind("\n")
                if last_newline > chunk_size // 2:
                    end = start + last_newline
                    chunk = text[start:end]
            chunks.append(chunk.strip())
            start = end - overlap
        return [c for c in chunks if c.strip()]


file_processor = FileProcessor()
